# server_cable_mcp.py
from __future__ import annotations

import re
import sys
import logging
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill

mcp = FastMCP("cable-request-mcp")

# stdout 오염 금지(중요): 로그는 stderr로만
logger = logging.getLogger("cable-mcp")
logger.setLevel(logging.INFO)
_handler = logging.StreamHandler(sys.stderr)
_handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
logger.addHandler(_handler)

OUT_DIR = Path("./artifacts").resolve()
OUT_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------
# Validation helpers
# -----------------------------
def _norm_port(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip()).lower()

def _norm(s: str) -> str:
    return (s or "").strip()

def _endpoint_key(row: Dict[str, Any], side: str) -> Tuple[str, str, str]:
    # side: "from" or "to"
    rack = _norm(row.get(f"{side}_rack", ""))
    dev = _norm(row.get(f"{side}_device", ""))
    port = _norm_port(row.get(f"{side}_port", ""))
    return (rack, dev, port)

def _cable_pair_key(row: Dict[str, Any]) -> Tuple[Tuple[str, str, str], Tuple[str, str, str], str]:
    a = _endpoint_key(row, "from")
    b = _endpoint_key(row, "to")
    ctype = _norm(row.get("cable_type", "")).upper()
    # A<->B 뒤집힌 중복을 잡기 위해 정렬
    if a <= b:
        return (a, b, ctype)
    return (b, a, ctype)

def _required_missing(row: Dict[str, Any]) -> List[str]:
    req = ["from_rack","from_device","from_port","to_rack","to_device","to_port","cable_type"]
    missing = [k for k in req if not _norm(str(row.get(k, "")))]
    return missing

# -----------------------------
# Excel template helpers
# -----------------------------
HEADER_SYNONYMS = {
    "from_rack": ["from rack", "출발 rack", "from_rack", "source rack", "rack a", "a rack", "시작 rack", "출발 랙", "from 랙"],
    "from_device": ["from device", "출발 device", "from_device", "source device", "device a", "a device", "출발 장비", "from 장비"],
    "from_port": ["from port", "출발 port", "from_port", "source port", "port a", "a port", "출발 포트", "from 포트"],
    "to_rack": ["to rack", "도착 rack", "to_rack", "dest rack", "rack b", "b rack", "도착 랙", "to 랙"],
    "to_device": ["to device", "도착 device", "to_device", "dest device", "device b", "b device", "도착 장비", "to 장비"],
    "to_port": ["to port", "도착 port", "to_port", "dest port", "port b", "b port", "도착 포트", "to 포트"],
    "cable_type": ["cable type", "type", "케이블 타입", "케이블type", "cable_type", "케이블 종류"],
    "qty": ["qty", "quantity", "수량", "ea", "개수"],
    "remark": ["remark", "note", "비고", "설명", "메모"],
}

def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _normalize_header(s: str) -> str:
    return re.sub(r"[\s\-_]+", " ", (s or "").strip().lower())

def _find_header_row(ws: Worksheet, scan_rows: int = 40) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    헤더 행을 찾아서:
      - header_row_index (1-based)
      - col_map: internal_key -> column_index(1-based)
    를 반환. 못 찾으면 None.
    """
    # 스캔하면서 “우리가 아는 헤더”가 여러 개 매칭되는 행을 찾음
    best = None
    best_score = 0

    for r in range(1, min(scan_rows, ws.max_row) + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        normed = [_normalize_header(_cell_text(v)) for v in row_values]

        col_map: Dict[str, int] = {}
        score = 0

        for key, syns in HEADER_SYNONYMS.items():
            syn_norms = {_normalize_header(x) for x in syns}
            for idx, header in enumerate(normed, start=1):
                if header in syn_norms and key not in col_map:
                    col_map[key] = idx
                    score += 1
                    break

        # 최소 핵심 7개(from/to/cable_type) 중 5개 이상 맞으면 후보
        core_keys = {"from_rack","from_device","from_port","to_rack","to_device","to_port","cable_type"}
        core_hit = len(core_keys.intersection(col_map.keys()))

        if core_hit >= 5 and score > best_score:
            best = (r, col_map)
            best_score = score

    return best

def _safe_out_xlsx(output_name: str) -> Path:
    name = Path(output_name).name
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    p = (OUT_DIR / name).resolve()
    if OUT_DIR not in p.parents:
        raise ValueError("Invalid output path")
    return p

def _write_rows(ws: Worksheet, header_row: int, col_map: Dict[str, int], rows: List[Dict[str, Any]]) -> int:
    """
    헤더 아래부터 rows를 쓴다.
    기존 데이터가 있으면 '아래에 append' (보수적으로)
    """
    start_row = header_row + 1
    # append row 찾기(핵심 컬럼 중 하나라도 값이 있는 마지막 행 아래)
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        any_val = False
        for k in ("from_device", "to_device", "from_port", "to_port"):
            ci = col_map.get(k)
            if ci and _cell_text(ws.cell(row=r, column=ci).value):
                any_val = True
                break
        if any_val:
            last = r
    write_row = last + 1

    # 스타일 템플릿: write_row가 header 바로 아래면 그 행을 기준, 아니면 last 행 기준
    style_row = max(header_row + 1, last) if ws.max_row >= (header_row + 1) else header_row

    def copy_style(src_r: int, dst_r: int, col: int):
        sc = ws.cell(row=src_r, column=col)
        dc = ws.cell(row=dst_r, column=col)
        dc._style = sc._style
        dc.number_format = sc.number_format
        dc.alignment = sc.alignment

    written = 0
    for row in rows:
        # 값 세팅
        for key, col in col_map.items():
            if key not in row:
                continue
            ws.cell(row=write_row, column=col).value = row[key]

        # 최소한의 보기 좋은 스타일 보정(템플릿 스타일이 있으면 복사)
        for key, col in col_map.items():
            copy_style(style_row, write_row, col)

        written += 1
        write_row += 1

    return written

# -----------------------------
# MCP Tools
# -----------------------------
@mcp.tool()
def cable_validate(connectivity_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    케이블 연결 리스트를 검증:
    - 필수값 누락
    - 동일 포트 중복 점유
    - A<->B 뒤집힌 중복
    """
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

    port_used: Dict[Tuple[str, str, str], int] = {}
    pair_seen: Dict[Tuple[Any, Any, str], int] = {}

    for i, row in enumerate(connectivity_rows):
        missing = _required_missing(row)
        if missing:
            errors.append({"row_index": i, "type": "MISSING_REQUIRED", "missing": missing})
            continue

        a = _endpoint_key(row, "from")
        b = _endpoint_key(row, "to")

        # 포트 충돌(같은 포트를 두 케이블이 사용)
        for ep, side in ((a, "from"), (b, "to")):
            if ep[0] == "" and ep[1] == "" and ep[2] == "":
                continue
            if ep in port_used:
                errors.append({
                    "row_index": i,
                    "type": "PORT_CONFLICT",
                    "message": f"Port already used: {side}={ep} (conflicts with row {port_used[ep]})",
                    "endpoint": {"side": side, "rack": ep[0], "device": ep[1], "port": ep[2]},
                    "conflicts_with_row": port_used[ep],
                })
            else:
                port_used[ep] = i

        # A<->B 중복
        pk = _cable_pair_key(row)
        if pk in pair_seen:
            warnings.append({
                "row_index": i,
                "type": "DUPLICATE_LINK",
                "message": f"Duplicate link (A<->B) detected (same as row {pair_seen[pk]})",
                "same_as_row": pair_seen[pk],
            })
        else:
            pair_seen[pk] = i

        # qty 기본값
        if "qty" in row:
            try:
                q = int(row["qty"])
                if q <= 0:
                    warnings.append({"row_index": i, "type": "QTY_NONPOSITIVE", "qty": row["qty"]})
            except Exception:
                warnings.append({"row_index": i, "type": "QTY_NOT_INT", "qty": row["qty"]})

    ok = len(errors) == 0
    summary = {
        "total_rows": len(connectivity_rows),
        "error_count": len(errors),
        "warning_count": len(warnings),
    }
    return {"ok": ok, "summary": summary, "errors": errors, "warnings": warnings}

@mcp.tool()
def cable_generate_request_excel(template_path: str, rows: List[Dict[str, Any]], output_name: str) -> Dict[str, Any]:
    """
    회사 템플릿 엑셀에 rows를 채워서 산출물 생성.
    - 템플릿에서 헤더 행/컬럼을 자동 탐지
    - 헤더 아래에 append
    """
    # 1) 먼저 검증
    v = cable_validate(rows)
    if not v["ok"]:
        return {"ok": False, "reason": "validation_failed", "validation": v}

    wb = load_workbook(template_path)
    # 우선 첫 시트부터 헤더 찾기 → 못 찾으면 전체 시트 스캔
    sheets = wb.worksheets

    found = None
    found_ws = None
    for ws in sheets:
        res = _find_header_row(ws)
        if res:
            found = res
            found_ws = ws
            break

    if not found or not found_ws:
        raise RuntimeError(
            "Could not find a header row in template. "
            "Please ensure the template has headers like From Rack/From Device/From Port/To Rack/To Device/To Port/Cable Type."
        )

    header_row, col_map = found
    ws = found_ws

    written = _write_rows(ws, header_row, col_map, rows)

    out_path = _safe_out_xlsx(output_name)
    wb.save(out_path)

    return {
        "ok": True,
        "path": str(out_path),
        "written_rows": written,
        "sheet": ws.title,
        "header_row": header_row,
        "mapped_columns": col_map,
        "validation": v,
    }

if __name__ == "__main__":
    mcp.run()
